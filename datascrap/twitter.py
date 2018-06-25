import requests as re
from bs4 import BeautifulSoup
import openpyxl
from requests.adapters import HTTPAdapter
import time

######################################################################
########Developer testing purpose
# raw_html_file = re.get('https://twitter.com/jsimini')
# # raw_html_file = re.get('https://twitter.com/eric_budzinski')
#
# print raw_html_file
# raw_html_file = raw_html_file.content
# # print raw_html_file
# soup_html = BeautifulSoup(raw_html_file, "html.parser")
# profile_div_html = soup_html.findAll("div", {"class": "ProfileSidebar ProfileSidebar--withLeftAlignment"})
#
# span = profile_div_html[0].findAll("span", {"class": "ProfileHeaderCard-urlText"})
# print len((span[0].get_text()).strip()),(span[0].get_text()).strip()
# link = ''
# if len((span[0].get_text()).strip())==0:
#     link = link
# else:
#     a = span[0].findAll("a")
#     link = (a[0].get('title')).strip()
# print link

# #
# # span = profile_div_html[0].findAll("span", {"class": "ProfileHeaderCard-locationText u-dir"})
# # print span[0].get_text()

################################################################################

def UI_path_URL_list_extract(excel_path, column_index):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.worksheets[0]
    UI_path_URL_list = list()

    for i in range(2,sheet.max_row ,1):
        UI_path_URL_list.append(sheet.cell(row=i, column=column_index).value)
    wb.close()
    return UI_path_URL_list

# def append_URL(excel_url):
#     start_url = 'https://app.lead411.com'
#     return start_url+excel_url

def location_name(profile_div):
    locname_span = profile_div[0].findAll("span", {"class": "ProfileHeaderCard-locationText u-dir"})
    raw_location_name =  locname_span[0].get_text()
    return raw_location_name.strip()

def web_link(profile_div):
    weblink_span = profile_div[0].findAll("span", {"class": "ProfileHeaderCard-urlText"})

    link = ''
    if len((weblink_span[0].get_text()).strip()) == 0:
        link = link
    else:
        a = weblink_span[0].findAll("a")
        link = (a[0].get('title')).strip()
    return link







column_index = 1 #give the column index which conatian urls
path = 'C:\Users\sjampala\Desktop\Book1.xlsx' #Give the path where excel file located

excel_urls = UI_path_URL_list_extract(path,column_index)

s = re.Session()
# s.headers['User-Agent'] = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.131 Safari/537.36'
# s.mount('http://', HTTPAdapter(max_retries=10))
s.mount('https://', HTTPAdapter(max_retries=10))


def return_div_html_url_data(url):

        try:
            _html_file = s.get(url)
        except:
            time.sleep(10)
        return _html_file

def div_html(raw__file):
    raw_html_ =  raw__file.content
    soup_html = BeautifulSoup(raw_html_, "html.parser")
    profile_div_html = soup_html.findAll("div", {"class": "ProfileSidebar ProfileSidebar--withLeftAlignment"})
    return  profile_div_html

for i in range(len(excel_urls)):
    rawfile = return_div_html_url_data(excel_urls[i])
    prfile_card_file = div_html(rawfile)
    location = location_name(prfile_card_file)
    weblink = web_link(prfile_card_file)
    print location,weblink
