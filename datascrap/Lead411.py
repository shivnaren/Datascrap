import requests as re
from bs4 import BeautifulSoup
import openpyxl
from requests.adapters import HTTPAdapter
import time

def UI_path_URL_list_extract(excel_path, column_index):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.worksheets[0]
    UI_path_URL_list = list()

    for i in range(2,sheet.max_row ,1):
        UI_path_URL_list.append(sheet.cell(row=i, column=column_index).value)
    wb.close()
    return UI_path_URL_list

def append_URL(excel_url):
    start_url = 'https://app.lead411.com'
    return start_url+excel_url

def return_a_html(raw_html_file):
    raw_html_content = raw_html_file.content
    soup_html = BeautifulSoup(raw_html_content, "html.parser")
    div_html = soup_html.findAll("div", {"class": "top_left_inner_box_url"})
    a_html = div_html[0].findAll('a')
    return a_html


column_index = 3
path = 'D:\Book1.xlsx'
links= list()
excel_urls = UI_path_URL_list_extract(path,column_index)
for i in range(len(excel_urls)):
    links.append(append_URL(excel_urls[i]))



s = re.Session()
s.headers['User-Agent'] = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.131 Safari/537.36'
s.mount('http://', HTTPAdapter(max_retries=10))
s.mount('https://', HTTPAdapter(max_retries=10))

def return_div_html_url_data(url):
        try:
            raw_html_file = s.get(url)
        except:
            time.sleep(5)
        return raw_html_file

for i in range(len(links)):
    html_f= return_div_html_url_data(links[i])
    a_html=  return_a_html(html_f)
    print a_html.get_text()


