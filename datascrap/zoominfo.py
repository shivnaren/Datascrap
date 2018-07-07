import requests as re
from bs4 import BeautifulSoup
import openpyxl
from requests.adapters import HTTPAdapter
import time



s = re.Session()
s.mount('http://', HTTPAdapter(max_retries=10))
s.mount('https://', HTTPAdapter(max_retries=10))



"""This function open the excel, xtract all links and close excel file"""
def UI_path_URL_list_extract(excel_path, column_index):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.worksheets[0]
    UI_path_URL_list = list()
    for i in range(2,sheet.max_row ,1):
        UI_path_URL_list.append(sheet.cell(row=i, column=column_index).value)
    wb.close()
    return UI_path_URL_list

"""This function open the link and return the raw html file
If the connection exceeds the more, then sleep for 10 sec.
"""
def return_div_html_url_data(url):
    try:
        html_file = s.get(url)
    except:
        time.sleep(10)
    return html_file

"""THis function prettify the raw html file data using Beautifulsoup"""
def div_html(raw__file):
    raw_html_ =  raw__file.content
    soup_html = BeautifulSoup(raw_html_, "html.parser")
    return soup_html


""" This funcftion checks whetehre phone data exts or not"""
def check_phone_div(souphtml):
    phone_div = souphtml.findAll("div", {"class":"content-box_row phone"})
    return phone_div

"""This function extract the phone data"""
def phone_data(phonediv):
    ph_span =  phonediv[0].findAll('span')
    return ph_span[0].get_text()

""" This funcftion checks whetehre web data exts or not"""
def check_web_link_exists(souphtml):
    wlink_div = souphtml.findAll("div", {"class":"content-box_row website"})
    return  wlink_div

"""This function extract the phone data"""
def weblink_data(webdiv):
    link = webdiv[0].findAll('a')
    return link[0].get_text()




column_index = 1 #give the column index which conatian urls
path = 'C:\Users\sjampala\Desktop\Book1.xlsx' #Give the path where excel file located

excel_urls = UI_path_URL_list_extract(path,column_index)

for link in excel_urls:
    raw_html_file = return_div_html_url_data(link)
    soup_html_file = div_html(raw_html_file)

    """Getting phone data"""
    check_phone_exists = check_phone_div(soup_html_file)
    if not check_phone_exists:
        continue
    ph_data = phone_data(check_phone_exists)


    """Getting Web link data"""
    check_web_exists = check_web_link_exists(soup_html_file)
    if not check_web_exists:
        continue
    wlink_data = weblink_data(check_web_exists)

    print ph_data, wlink_data
