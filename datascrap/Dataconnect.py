import requests as re
from bs4 import BeautifulSoup
import openpyxl
from requests.adapters import HTTPAdapter
import time




def UI_path_URL_list_extract(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.worksheets[0]
    UI_path_URL_list = list()

    for i in range(2,sheet.max_row ,1):
        UI_path_URL_list.append(sheet.cell(row=i, column=column_index).value)
    wb.close()
    return UI_path_URL_list


# raw_html_file = re.get('https://connect.data.com/company/view/vNTDYqDD4A0r9ybX1QtHXg/')

# raw_html_content = raw_html_file.content
# soup_html = BeautifulSoup(raw_html_content, "html.parser")
# # print(soup_html.prettify())
#
# div_html = soup_html.findAll("div", {"class":"seo-company-info"})
# # tables = div_html[0].findAll('table')

def return_div_html(url):
    raw_html_file = re.get(url)
    raw_html_content = raw_html_file.content
    soup_html = BeautifulSoup(raw_html_content, "html.parser")
    div_html = soup_html.findAll("div", {"class": "seo-company-info"})

def append_URL(excel_url):
    start_url = 'https://connect.data.com'
    return start_url+excel_url


# print(append_URL('company/view/vNTDYqDD4A0r9ybX1QtHXg/'))

#
def find_tables_html(raw_div_html):
    for i in range(len(raw_div_html)):
        tables_html = raw_div_html[i].findAll('table')
    return tables_html

def find_rows_html(raw_tables_html):
    for i in range(len(raw_tables_html)):
        row_html = raw_tables_html[i].findAll('tr')
    return row_html

def find_name_row_html(raw_row_html):
    for i in range(len(raw_row_html)):
        row_data = raw_row_html[i].get_text()
        if "Name" in row_data:
            row_name_html = raw_row_html[i].find_all('td')
    return row_name_html

def find_name_td_html(raw_row_name_html):
    name_list = list()
    for i in range(len(raw_row_name_html)):
        name_list.append(raw_row_name_html[i].get_text())
    return name_list

def find_website_row_html(raw_row_html):
    for i in range(len(raw_row_html)):
        row_data = raw_row_html[i].get_text()
        if "Website" in row_data:
            row_website_html = raw_row_html[i].find_all('td')
    return row_website_html

def find_website_td_html(raw_row_name_html):
    website_list = list()
    for i in range(len(raw_row_name_html)):
        website_list.append(raw_row_name_html[i].get_text())
    return website_list




def find_phone_row_html(raw_row_html):
    for i in range(len(raw_row_html)):
        row_data = raw_row_html[i].get_text()
        if "Phone" in row_data:
            row_phone_html = raw_row_html[i].find_all('td')
    return row_phone_html

def find_phone_td_html(raw_row_phone_html):
    phone_list = list()
    for i in range(len(raw_row_phone_html)):
        phone_list.append(raw_row_phone_html[i].get_text())
    return phone_list

# tables_html = find_tables_html(div_html)
# rows_html = find_rows_html(tables_html)
# phone_row_html = find_phone_row_html(rows_html)
# phone_data = find_phone_td_html(phone_row_html)
#
# name_row_html = find_name_row_html(rows_html)
# name_data = find_name_td_html(name_row_html)
#
# web_site_html = find_website_row_html(rows_html)
# website_data = find_website_td_html(web_site_html)
#
#
# print(phone_data)
# print(name_data)
# print(website_data)
#



links= list()
path = 'D:\Book1.xlsx'
column_index = 3

excel_urls = UI_path_URL_list_extract(path)
for i in range(len(excel_urls)):

    links.append(append_URL(excel_urls[i]))



s = re.Session()
s.headers['User-Agent'] = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.131 Safari/537.36'
s.mount('http://', HTTPAdapter(max_retries=10))
s.mount('https://', HTTPAdapter(max_retries=10))

def return_div_html_url_data(url):
        try:
            raw_html_file = s.get(url)
        except:
            time.sleep(5)
        return raw_html_file


def raw_div_html(raw_html_file):
        raw_html_content = raw_html_file.content
        soup_html = BeautifulSoup(raw_html_content, "html.parser")
        div_html = soup_html.findAll("div", {"class": "seo-company-info"})
        return div_html








for i in range(len(links)):
    html_f= return_div_html_url_data(links[i])

    tables_html = find_tables_html(raw_div_html(html_f))
    rows_html = find_rows_html(tables_html)

    phone_row_html = find_phone_row_html(rows_html)
    phone_data = find_phone_td_html(phone_row_html)

    name_row_html = find_name_row_html(rows_html)
    name_data = find_name_td_html(name_row_html)

    web_site_html = find_website_row_html(rows_html)
    website_data = find_website_td_html(web_site_html)

    print(str(phone_data) + '------------'+str(name_data)+ '------------'+str(website_data))


# # # print(tables)
# # #
# # for i in range(len(tables)):
# #     tr = tables[i].find_all('tr')
# #     for i in range(len(tr)):
# #         td = tr[i].find_all('td')
# #         for i in range(len(td)):
# #             data = td[i].get_text()
# #             print(data)
#
#
# #
#
#
# #
# #
# #
#
# # for i in range(len(tables)):
# #     tr = tables[i].find_all('tr')
# #     for i in range(len(tr)):
# #         data = tr[i].get_text()
# #         if "Phone" in data:
# #             td = tr[i].find_all('td')
#
#
#
#
# #
# #
# #
